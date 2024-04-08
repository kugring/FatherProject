import os
import re
import pyautogui
import pyperclip
from selenium.common.exceptions import NoSuchElementException
import time
from selenium.webdriver.common.by import By
from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# 파일 폴더 자리
folder_path = r"C:\Users\gram\Desktop\7. 복흥면\4. 율평마을 농로 선형개선공사 1000"

# 변수 초기화
quantity_file_path = None

# 폴더 내의 모든 파일을 확인하여 "수량 -"이라는 단어가 들어가는 엑셀 파일을 찾음
for file in os.listdir(folder_path):
    if file.endswith('.xlsx') and "수량-" in file:
        quantity_file_path = os.path.join(folder_path, file)
        break
else:
    print("폴더 내에 '수량-'이라는 단어가 들어가는 엑셀 파일이 없습니다.")
    exit()

# 수량이라는 단어가 들어간 엑셀 파일 경로 출력
print("수량이라는 단어가 들어간 엑셀 파일의 경로:", quantity_file_path)


# 엑셀 파일 내 "용지조서(test)" 시트에서 데이터를 읽어오는 함수
def read_data_from_excel(file_path, sheet_name, start_row, start_col):
    # 엑셀 파일 열기
    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]

    # 시작 셀부터 데이터 읽기
    data = []
    row_index = start_row
    col_index = start_col
    while True:
        cell_value = ws.cell(row=row_index, column=col_index).value
        if cell_value:
            data.append(cell_value)
            row_index += 1
        else:
            break

    # 엑셀 파일 닫기
    wb.close()
    return data


# "용지조서(test)" 시트의 데이터를 읽어옵니다.
data = read_data_from_excel(quantity_file_path, "용지조서(test)", start_row=5, start_col=2)

texts = []
for value in data:
    print(value)
    texts.append(value)

# 단어들을 담은 리스트 출력
print("출력한 단어들을 리스트에 담았습니다:", texts)

# 웹 드라이버 생성
driver = webdriver.Edge()

# pdf파일 다른이름으로 저장시 필요함
folder_name = folder_path + "\\토지대장\\"

# 크롬 웹 드라이버 경로 설정 (본인 컴퓨터에 설치된 드라이버 경로로 설정해주세요)
edge_driver_path = r"C:\Users\gram\Desktop\edgedriver_win64\msedgedriver.exe"

# Edge 웹 드라이버 옵션 설정
options = webdriver.EdgeOptions()
options.add_argument("start-maximized")  # 창 최대화

# 웹 드라이버 실행
driver = webdriver.Edge(options=options)


def element_click(driver, xpath):
    while True:
        try:
            # 발급하기 버튼 클릭
            element = driver.find_element(By.XPATH, f'{xpath}')
            element.click()
            break  # 요소를 찾았으므로 루프 종료
        except NoSuchElementException:
            # 요소가 발견되지 않으면 잠시 대기하고 다시 시도
            time.sleep(1)  # 1초 대기 후 다시 시도


# 해당 사이트로 이동
driver.get("https://www.gov.kr/mw/AA020InfoCappView.do?CappBizCD=13100000026&HighCtgCD=A02001001&tp_seq=01&Mcode=10207")

time.sleep(1)

# '발급하기' 버튼 누르는 코드
element_click(driver, '//*[@id="applyBtn"]/a')
time.sleep(1)

# '회원 신청하기' 버튼 누르는 코드
element_click(driver, '//*[@id="memberApplyBtn"]')

# 로그인창에서 아이디페이지를 선택하는 코드
element = driver.find_element(By.XPATH, '//*[@id="아이디"]')
element.click()

# 로그인창에서 아이디 인풋을 선택하는 코드
element = driver.find_element(By.XPATH, '//*[@id="userId"]')
element.send_keys('wogur824333')

# 다음버튼을 누르는 코드
element = driver.find_element(By.XPATH, '//*[@id="genLogin"]')
element.click()

# 비밀번호 입력하는 코드
element = driver.find_element(By.XPATH, '//*[@id="pwd"]')
element.send_keys('wogurdl8243@')

# 로그인 버튼 누르는 코드
element = driver.find_element(By.XPATH, '//*[@id="genLogin"]')
element.click()

for text in texts:

    # 4자리 부터 끝까지 '산'이라는 단어가 들어가는 확인하는 if문
    substring = text[4:]
    if '산' in substring:
        # 문자열에서 뒤에서부터 "산"이 나오는 인덱스 찾기
        index_of_san = text.rfind('산')

        # "산" 이후의 문자열을 추출하고, "산" 문자 제거
        text = text[:index_of_san] + text[index_of_san + 1:]
        임야 = True

    # 문자, 숫자(하이픈 포함), 나머지 패턴 정의
    pattern = re.compile(r'([가-힣]+)\s+((\d+(-\d+)?))')

    # 정규식을 사용하여 문자열을 나눔
    match = pattern.match(text)
    if match:
        # 각 파트 출력
        토지명 = match.group(1)  # 첫 번째 그룹: 문자 또는 숫자
        번지수 = match.group(2)  # 두 번째 그룹: 숫자 또는 외의 표현
        번지호 = ""  # 세 번째 파트 초기화

        # 두 번째 그룹이 숫자(또는 하이픈을 포함한 숫자)인 경우
        if re.match(r'\d', 번지수):
            # 숫자와 나머지로 분리
            번지수_match = re.match(r'(\d+)(.*)', 번지수)
            번지수 = 번지수_match.group(1)  # 두 번째 파트: 숫자
            번지호 = 번지수_match.group(2).replace("외", "").strip()  # 세 번째 파트: 나머지 문자열, '외' 및 공백 제거
            번지호 = 번지수_match.group(2).replace("-", "").strip()  # 세 번째 파트: 나머지 문자열, '-'와 '산' 제거

    else:
        print("일치하는 패턴이 없습니다.")

    ##########################################################################################

    ''' 여기서 (토지) or (산) 이냐에 따라서 if문을 작성하면 된다'''

    # '산'이라는 단어가 포함되어 있는지 확인
    try:
        if 임야:
            # (대장 구분)_임야대장 선택
            element_click(driver, '// *[ @ id = "main"] / div / div[1] / div / div[2] / label')
            임야 = False
        else:
            # (대장 구분)_토지대장 선택
            element_click(driver, '// *[ @ id = "main"] / div / div[1] / div / div[1] / label')
    except:
        time.sleep(0.1)
    ##########################################################################################

    # 원래 창의 핸들을 저장
    original_window_handle = driver.current_window_handle

    # 주소검색하는 클릭하는 코드
    element_click(driver, '//*[@id="btnAddress"]')

    time.sleep(1)

    # 새로운 창 또는 모달 창으로 전환
    # 모든 창 핸들 가져오기
    all_window_handles = driver.window_handles
    time.sleep(1)

    # 새로운 창 핸들 찾기
    new_window_handle = None
    for handle in all_window_handles:
        if handle != original_window_handle:
            new_window_handle = handle
            break
    time.sleep(3)

    while True:
        try:
            # 새로운 창 핸들로 전환
            driver.switch_to.window(new_window_handle)
            break
        except:
            time.sleep(1)

    # 주소창에 글자를 입력하는 코드
    element = driver.find_element(By.XPATH, '//*[@id="txtAddr"]')
    element.send_keys(토지명 + 번지수)

    # 주소 검색버튼을 누르는 코드
    element = driver.find_element(By.XPATH, '//*[@id="frm_popup"]/fieldset/div/div/span/button')
    element.click()

    try:
        # address-result-list 클래스를 가진 요소 찾기
        address_list = driver.find_element(By.CLASS_NAME, "address-result-list")

        # address-result-list의 자식 요소들 중에서 span 태그의 텍스트를 확인하고 "전북특별자치도 순창군"이 있는지 확인
        elements = address_list.find_elements(By.XPATH, './/div')
        for element in elements:
            if "전북특별자치도 순창군" in element.text:
                # 찾았으면 해당 요소를 클릭
                element.click()
                break
        else:
            print("찾을 수 없습니다.")

    except Exception as e:
        print("오류 발생:", e)

    # 기존 창으로 다시 전환
    driver.switch_to.window(original_window_handle)

    # 앞자리 번지수
    element = driver.find_element(By.XPATH, '//*[@id="토지임야대장신청서_IN-토지임야대장신청서_신청토지소재지_주소정보_상세주소_번지"]')
    element.clear()
    element.send_keys(번지수)

    if 번지호:
        # 뒷자리 번지수
        element = driver.find_element(By.XPATH, '//*[@id="토지임야대장신청서_IN-토지임야대장신청서_신청토지소재지_주소정보_상세주소_호"]')
        element.clear()
        element.send_keys(번지호)

    # 신청하기 버튼 클릭
    element = driver.find_element(By.XPATH, '//*[@id="btn_end"]')
    element.click()

    # 테이블표에서 발급하기 버튼 클릭
    element_click(driver, '/html/body/div[7]/div[2]/div[2]/div/div/div[1]/table/tbody/tr[1]/td[4]/p[2]/span/a')

    # 모든 창 핸들 가져오기
    all_window_handles = driver.window_handles

    # 새로운 창 핸들 찾기
    new_window_handle = None
    for handle in all_window_handles:
        if handle != original_window_handle:
            new_window_handle = handle
            break
    # 새로운 창 핸들로 전환
    driver.switch_to.window(new_window_handle)

    ######################################################################################################################
    while True:
        try:
            # iframe 요소 식별
            iframe = driver.find_element(By.ID, "viewerFrame")
            break
        except:
            time.sleep(1)

    # iframe으로 전환
    driver.switch_to.frame(iframe)

    while True:
        try:
            # 해당 요소 찾기
            element = driver.find_element(By.XPATH, '//*[@id="pageContainer1"]/div[2]')
            break
        except:
            time.sleep(1)

    time.sleep(5)
    # 해당 요소의 내부 HTML 가져오기
    html_content = element.get_attribute('innerHTML')

    # BeautifulSoup을 사용하여 HTML 파싱
    soup = BeautifulSoup(html_content, 'html.parser')

    # '등록 번호'를 포함하는 첫 번째 div 태그 찾기
    target_text = '등 록 번 호'
    div_tag = soup.find('div', string=target_text)

    # '등록 번호'를 포함하는 div 태그의 순서 확인
    if div_tag:
        index = 1
        sibling = div_tag.find_previous_sibling()
        while sibling:
            if sibling.name == 'div':
                index += 1
            sibling = sibling.find_previous_sibling()
    # 내부 페이지의 소스 가져오기
    inner_page_source = driver.page_source
    주소 = driver.find_element(By.XPATH, f'//*[@id="pageContainer1"]/div[2]/div[{index + 2}]')
    소유주 = driver.find_element(By.XPATH, f'//*[@id="pageContainer1"]/div[2]/div[{index + 5}]')
    지번 = driver.find_element(By.XPATH, f'//*[@id="pageContainer1"]/div[2]/div[27]')
    print('주  소:', 주소.text)
    print('소유주:', 소유주.text)
    print('지  번:', 지번.text)

    # 다시 메인 프레임으로 전환
    driver.switch_to.default_content()

    ######################################################################################################################

    # 인쇄버튼 누르기
    element_click(driver, '/html/body/div[6]/header/div[2]/div[3]/ul/li')

    time.sleep(1)
    # 인쇄대화상자가 뜨는데 엔터를 누른다.
    pyautogui.press("enter")

    time.sleep(2)
    # 키보드로 "토지명" 입력
    pyperclip.copy(folder_name + text)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(1)

    # 파일저장하기 (엔터키 누르기)
    pyautogui.press("enter")
    time.sleep(1)

    # 현재 창 닫기
    driver.close()

    # 기존 창으로 다시 전환
    driver.switch_to.window(original_window_handle)

    # 설문조사 등장시 닫기 누르는 코드
    try:
        element = driver.find_element(By.XPATH, '/html/body/div[1]/div/button')
        element.click()
        alert = driver.switch_to.alert
        alert.accept()  # "Yes" 또는 "OK" 버튼 클릭
    except NoSuchElementException:
        time.sleep(1)

    # 기존 창으로 다시 전환
    driver.switch_to.window(original_window_handle)

    # 추가신청을 진행한다.
    try:
        element = driver.find_element(By.XPATH,'/html/body/div[8]/div[2]/div[2]/div/div/div[1]/table/tbody/tr[2]/td[7]/span/a')
        element.click()
    except:
        element = driver.find_element(By.XPATH,'/html/body/div[7]/div[2]/div[2]/div/div/div[1]/table/tbody/tr[1]/td[7]/span/a')
        element.click()


    # 추기선청 alert창을 승인하는 코드
    alert = driver.switch_to.alert
    alert.accept()  # "Yes" 또는 "OK" 버튼 클릭

while True: pass
