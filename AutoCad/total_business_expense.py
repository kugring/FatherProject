import xlrd
from openpyxl import load_workbook
import os


# 폴더 경로
folder_path = r'C:\Users\gram\Desktop\7. 복흥면\4. 율평마을 농로 선형개선공사 1000'


def 갑지_총사업비(folder_path):
    # 폴더 내의 모든 파일 확인하여 xls 파일 찾기
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xls'):
            ESTX_file_path = os.path.join(folder_path, file_name)
            break
    else:
        print("폴더 내에 xls 파일이 없습니다.")
    # 폴더 내의 모든 파일을 확인하여 "갑지-"이라는 단어가 들어가는 엑셀 파일을 찾음
    for file in os.listdir(folder_path):
        if file.endswith('.xlsx') and "갑지-" in file:
            title_excel_path = os.path.join(folder_path, file)
            break
    else:
        source_file_path = r"C:\Users\gram\Desktop\갑지- (견본).xlsx"

        # 목표 파일 경로 생성
        title_excel_path = os.path.join(folder_path, os.path.basename(source_file_path))

        # 파일 복사
        shutil.copy(source_file_path, title_excel_path)
        print("폴더 내에 '갑지'파일이 없습니다. 새로 생성합니다.")

    # 파일명 추출
    file_name_with_extension = os.path.basename(ESTX_file_path)
    file_name = os.path.splitext(file_name_with_extension)[0]
    # 엑셀 파일 열기
    workbook = xlrd.open_workbook(ESTX_file_path)

    # 원가계산서 시트 선택
    sheet = workbook.sheet_by_name('원가계산서')

    # 탐색할 단어들과 변수명 딕셔너리 생성
    target_words = {
        "총     공    사     비": "total_cost",
        "관   급   자   재   대": "government_aid",
        "도        급        액": "provincial_aid",
        "부   가   가   치   세": "overhead_tax",
        "총        원        가": "total_price"
    }

    # 각 단어를 변수명으로 사용하여 값을 저장할 변수들을 동적으로 생성
    for variable_name in target_words.values():
        globals()[variable_name] = None  # 변수를 None으로 초기화

    # 엑셀 시트를 순회하면서 각 단어를 찾음
    for target_word, variable_name in target_words.items():
        for row_index in range(sheet.nrows):
            for col_index in range(sheet.ncols):
                cell_value = sheet.cell_value(row_index, col_index)
                if cell_value == target_word:
                    # 발견된 셀의 오른쪽에 있는 값을 가져옴
                    if col_index + 1 < sheet.ncols:  # 열 범위를 넘어가지 않도록 체크
                        next_cell_value = sheet.cell_value(row_index, col_index + 2)
                        # 빈 문자열인 경우 0으로 처리하고 정수로 변환
                        next_cell_value = next_cell_value if next_cell_value else '0'
                        # 값을 해당 변수에 할당
                        globals()[variable_name] = int(float(next_cell_value))

    # 엑셀 파일 닫기
    workbook.release_resources()

    # 엑셀 파일을 쓰기 모드로 열기
    wb = load_workbook(filename=title_excel_path)
    ws = wb['갑지']

    # 엑셀 시트를 순회하면서 각 단어를 찾고 값을 입력
    for target_word, variable_name in target_words.items():
        for row_index in range(1, ws.max_row + 1):
            for col_index in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_index, column=col_index).value
                if cell_value == target_word:
                    # 발견된 셀의 오른쪽에 변수에 저장된 값을 넣음
                    ws.cell(row=row_index, column=col_index + 1).value = globals()[variable_name]
                    print(target_word, ":", globals()[variable_name])

    # 현재 스크립트 파일의 경로
    script_directory = os.path.dirname(os.path.abspath(ESTX_file_path))

    # 새로운 파일 경로 (현재 스크립트 파일의 위치와 같은 곳에 저장)
    new_file_path = os.path.join(script_directory, f"갑지-{file_name}.xlsx")

    # 변경된 엑셀 파일 저장
    wb.save(filename=new_file_path)
    print("---------------------------------------------------------------")
