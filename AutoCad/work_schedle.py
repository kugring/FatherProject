import xlrd
from openpyxl import load_workbook
import os

# 폴더 경로
folder_path = r'C:\Users\gram\Desktop\7. 복흥면\4. 율평마을 농로 선형개선공사 1000'


def 갑지_예정공정표(folder_path):
    '''
     *주의사항*
        - 비용합계는 'U열"에 있을것!
        - 실행 이후에 열 숨기기 가능함
        - 'B열'은 5개만 있으므로 5개만 가능 (* 추후 B열의 행을 늘리면 가능함 *)
    '''

    # 폴더 내의 모든 파일 확인하여 xls 파일 찾기
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xls'):
            ESTX_file_path = os.path.join(folder_path, file_name)
            break
    else:
        print("폴더 내에 xls 파일이 없습니다.")

    # 엑셀 파일 열기
    workbook = xlrd.open_workbook(ESTX_file_path)

    # "내역서총괄표" 시트 선택
    sheet = workbook.sheet_by_name('내역서총괄표')

    # 변수 초기화
    variables = {}
    # 공사명들
    공사명들 = []
    # 공사합계
    공사합계 = []
    # B열에 값이 있는 경우에만 해당 행의 B열 값과 오른쪽에 있는 1번째 열, 6번째 열의 값을 출력하고 변수에 저장
    for row_index in range(3, sheet.nrows):  # 0-based index이므로 B4는 (3, 1)입니다.
        cell_value_b = sheet.cell_value(row_index, 1)  # B열의 인덱스는 1입니다.
        if cell_value_b and isinstance(cell_value_b, str):
            variable_name = str(sheet.cell_value(row_index, 2))  # B열의 오른쪽 1번째 열의 인덱스는 2입니다.
            공사명들.append(variable_name)
            variable_value = str(sheet.cell_value(row_index, 6))[:-5]  # B열의 오른쪽 6번째 열의 인덱스는 6입니다.
            공사합계.append(variable_value)
            variables[variable_name] = variable_value
    # 사급자재대는 제외, 토공사를 '토공'으로 표기한다.
    공사명들 = ['토공' if 공사명 == '토공사' else 공사명 for 공사명 in 공사명들 if 공사명 != '사급자재대*']
    공사명들 = [' '.join(공사명) for 공사명 in 공사명들]

    # 엑셀 파일 닫기
    workbook.release_resources()

    # 폴더 내의 모든 파일을 확인하여 "수량 -"이라는 단어가 들어가는 엑셀 파일을 찾음
    for file in os.listdir(folder_path):
        if file.endswith('.xlsx') and "갑지-" in file:
            title_excel_path = os.path.join(folder_path, file)
            break
    else:
        # 현재 스크립트 파일의 경로
        script_directory = os.path.dirname(os.path.abspath(ESTX_file_path))
        title_excel_path = os.path.join(script_directory, f'갑지-{file_name}.xlsx')
        print("폴더 내에 '갑지'파일이 없습니다. 새로 생성합니다.")

    # 엑셀 파일에 데이터 입력하기
    def set_data_to_excel(file_path, sheet_name, cell, data):
        wb = load_workbook(filename=file_path)
        ws = wb[sheet_name]
        ws[cell] = data
        wb.save(filename=file_path)
        wb.close()

    # indexes 리스트를 생성합니다.
    indexes = [5 + 3 * i for i in range(len(공사명들))]

    # indexes 리스트의 각 인덱스에 대해 작업을 수행합니다.
    for idx, index in enumerate(indexes):
        # 변수명을 가져와서 엑셀 파일에 데이터를 입력합니다.
        set_data_to_excel(title_excel_path, '예정공정.동원인원', f'U{index}', int(공사합계[idx]))

    # indexes 리스트의 각 인덱스에 대해 작업을 수행합니다.
    for idx, index in enumerate(indexes):
        # 변수명을 가져와서 엑셀 파일에 데이터를 입력합니다.
        set_data_to_excel(title_excel_path, '예정공정.동원인원', f'B{index}', 공사명들[idx])

    print("예정공정표 출력완료!")
    print("---------------------------------------------------------------")
