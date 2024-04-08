import os
from openpyxl import load_workbook
from shutil import copyfile
from PyQt5.QtWidgets import QFileDialog

def create_folder(base_path):
    # 찾을 파일의 키워드
    keyword = "실시설계용역"

    # 파일 경로를 저장할 변수 초기화
    found_file_path = None

    # 기본 경로 안의 모든 파일을 탐색
    for root, dirs, files in os.walk(base_path):
        for file_name in files:
            if file_name.endswith(".xlsm") and keyword in file_name:
                found_file_path = os.path.join(root, file_name)
                break
        if found_file_path:
            break

    # 결과 출력
    if found_file_path:
        print("찾은 파일 경로:", found_file_path)
        # 파일이 위치한 상위 상위 폴더 경로 가져오기
        parent_parent_folder = os.path.abspath(os.path.join(found_file_path, os.pardir, os.pardir))  # 상위 상위 폴더의 경로

        # 엑셀 파일 열기
        workbook = load_workbook(found_file_path)
        sheet = workbook.active

        # 시작 셀과 끝 셀 설정
        start_cell = "B4"
        end_cell = "J4"

        # 시작 셀과 끝 셀 좌표 구하기
        start_column = ord(start_cell[0]) - 64  # 엑셀 열의 알파벳을 숫자로 변환
        start_row = int(start_cell[1:])
        end_column = ord(end_cell[0]) - 64
        end_row = int(end_cell[1:])

        # 찾을 단어 리스트
        target_words = ['연번', '사 업 명(부기)', '사업비']

        # 시작 셀부터 끝 셀까지 순회하면서 target_words를 포함하는 열 찾기
        target_columns = []
        for col in range(start_column, end_column + 1):
            cell_value = sheet.cell(row=start_row, column=col).value
            if cell_value in target_words and cell_value != '사업개요':  # '사업개요'는 제외
                target_columns.append(col)

        # 폴더 생성
        max_row = sheet.max_row
        for row in sheet.iter_rows(min_row=7, max_row=max_row - 1, min_col=min(target_columns),
                                   max_col=max(target_columns)):
            formatted_data = ''
            for idx, cell in enumerate(row):
                print(idx, cell)
                if idx + start_column in target_columns:
                    if cell.value is None or cell.value == '':  # 빈 문자열인 경우 처리
                        value = ''
                    elif isinstance(cell.value, int) and cell.value >= 10000:  # 10000 이상인 경우에만 처리
                        str_value = str(cell.value)
                        if str_value.endswith('0'):
                            value = int(str_value[:-4])
                        else:
                            value = cell.value
                    else:
                        value = cell.value
                    # 연번과 사업 명 등 데이터 추가
                    if idx + start_column == target_columns[0]:
                        formatted_data += f'{value}.'
                    elif idx + start_column == target_columns[1]:
                        formatted_data += f' {value}'
                    elif idx + start_column == target_columns[2]:
                        formatted_data += f' {value}'

            # 폴더명 생성
            folder_name = formatted_data.strip()  # 폴더명에서 앞뒤 공백 제거
            folder_path = os.path.join(parent_parent_folder, folder_name)
            os.makedirs(folder_path, exist_ok=True)  # 폴더가 없으면 생성하고, 있으면 pass

            # "토지대장X" 폴더 생성
            subfolder_path = os.path.join(folder_path, "X토지대장")  # 토지대장X 폴더 경로 수정
            if not os.path.exists(subfolder_path):  # 폴더가 없을 때만 생성
                os.makedirs(subfolder_path)  # 폴더 생성

            # "(견본)"이 포함된 파일을 찾아서 해당 폴더로 복사
            keyword = "(견본)"
            for root, dirs, files in os.walk(base_path):
                for file_name in files:
                    if keyword in file_name:
                        sample_file_path = os.path.join(root, file_name)
                        # 새로운 파일 이름 생성
                        new_file_name = formatted_data.strip()  # 폴더명에서 앞뒤 공백 제거
                        print(new_file_name)  # 디버깅용 출력
                        # 숫자를 제거한 뒤 "갑지-"와 합치기
                        new_file_name = "X갑지- {}.xlsm".format("".join(filter(str.isalpha, new_file_name)))
                        # 새로운 파일 경로 생성
                        target_file_path = os.path.join(folder_path, new_file_name)
                        # 해당 파일을 새로운 이름으로 폴더로 복사
                        copyfile(sample_file_path, target_file_path)
                        print(f"'{file_name}' 파일이 폴더 '{folder_path}'에 '{new_file_name}'로 복사되었습니다.")  # 출력 메시지 수정

        print("폴더 생성 및 파일 복사가 완료되었습니다.")
    else:
        print("해당 조건을 만족하는 파일을 찾지 못했습니다.")


