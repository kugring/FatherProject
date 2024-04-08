from openpyxl.worksheet.page import PageMargins
from openpyxl import load_workbook
from openpyxl.worksheet.page import PrintOptions

'''
    - 참조(전),위치조서(까지)_프린트가 진행되는 코드
    - 여백을 자동으로 진행시켜주는 코드
'''


def print_until_sheet(file_path, stop_sheet_name):
    wb = load_workbook(file_path)
    sheets = wb.sheetnames

    for sheet_name in sheets:
        ws = wb[sheet_name]
        ws.sheet_state = 'visible'  # 시트를 보이도록 설정합니다.

        # 프린트 옵션 설정
        ws.print_options = PrintOptions(horizontalCentered=True)

        # 여백 설정 (좁은 여백)
        ws.page_margins = PageMargins(left=0.25197, right=0.25197, top=0.7512, bottom=0.7512, header=0.299,
                                      footer=0.299)

        if sheet_name == stop_sheet_name:
            break

        if sheet_name == "위치참조":
            ws.sheet_state = 'visible'  # 시트를 보이도록 설정합니다.
            ws.print_area = ws.dimensions  # 시트의 인쇄 영역을 시트의 모든 영역으로 설정합니다.
            wb.save(file_path)
            ws.print_out()  # 시트를 프린트합니다.
            break

        elif sheet_name != "위치참조":  # 위치참조 이후부터 프린트합니다.
            ws.sheet_state = 'visible'  # 시트를 보이도록 설정합니다.
            ws.print_area = ws.dimensions  # 시트의 인쇄 영역을 시트의 모든 영역으로 설정합니다.
            wb.save(file_path)
            ws.print_out()  # 시트를 프린트합니다.


file_path = "example.xlsx"  # 엑셀 파일 경로
stop_sheet_name = "참조"  # 프린트를 멈출 시트 이름

print_until_sheet(file_path, stop_sheet_name)
