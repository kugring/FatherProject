from openpyxl import load_workbook
import os
import shutil

'''

    *사용방법*
     - 엑셀(수량)파일에서 ['공사명','공사위치']를 추출해서
       엑셀(갑지)파일의 ['사업개요']에 저장한다.

    *주의사항*
    - 수량산출서 파일명을 "수량- "으로 정확히 앞에 기입해줘야한다.
'''

folder_path = r"C:\Users\gram\Desktop\7. 복흥면\4. 율평마을 농로 선형개선공사 1000"


def 갑지_공사명_공사위치(folder_path):
    print("---------------------------------------------------------------")

    # 엑셀 파일에서 데이터 가져오기
    def get_data_from_excel(file_path, sheet_name, cell):
        wb = load_workbook(filename=file_path)
        ws = wb[sheet_name]
        data = ws[cell].value
        wb.close()
        return data

    # 엑셀 파일에 데이터 입력하기
    def set_data_to_excel(file_path, sheet_name, cell, data):
        wb = load_workbook(filename=file_path)
        ws = wb[sheet_name]
        ws[cell] = data
        wb.save(filename=file_path)
        wb.close()

    # 폴더 내의 모든 파일을 확인하여 "수량 -"이라는 단어가 들어가는 엑셀 파일을 찾음
    for file in os.listdir(folder_path):
        if file.endswith('.xlsx') and "수량-" in file:
            quantity_file_path = os.path.join(folder_path, file)
            # "수량- "을 제거한 파일명 가져오기
            공사명, _ = os.path.splitext(file.replace("수량- ", ""))
            break
    else:

        print("폴더 내에 '수량 -'이라는 단어가 들어가는 엑셀 파일이 없습니다.")
        exit()

    # 폴더 내의 모든 파일을 확인하여 "갑지 -"이라는 단어가 들어가는 엑셀 파일을 찾음
    for file in os.listdir(folder_path):
        if file.endswith('.xlsx') and "갑지-" in file:
            title_excel_path = os.path.join(folder_path, file)
            break
    else:
        source_file_path = r"C:\Users\gram\Desktop\갑지- (견본).xlsx"

        # 목표 파일 경로 생성
        target_file_path = os.path.join(folder_path, os.path.basename(source_file_path))

        # 파일 복사
        shutil.copy(source_file_path, target_file_path)
        print("폴더 내에 '갑지'파일이 없습니다. 새로 생성합니다.")

    # 수량 파일에서 데이터 가져오기
    공사위치 = get_data_from_excel(quantity_file_path, '용지조서', 'B2')
    # 갑지 파일에 공사위치 데이터 입력하기
    set_data_to_excel(title_excel_path, '사업개요', 'C2', 공사위치)
    print(f'공사위치:[{공사위치}] - (입력완료)')
    # 갑지 파일에 공사명 데이터 입력하기
    set_data_to_excel(title_excel_path, '사업개요', 'C1', 공사명)
    print(f'공사명:[{공사명}] - (입력완료)')
    print("---------------------------------------------------------------")
