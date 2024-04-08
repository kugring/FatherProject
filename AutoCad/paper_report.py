from openpyxl import load_workbook
import os

# 파일 폴더 경로
folder_path = r"C:\Users\gram\Desktop\7. 복흥면\4. 율평마을 농로 선형개선공사 1000"

# 변수 초기화
quantity_file_path = None

# 폴더 내의 모든 파일을 확인하여 "수량 -"이라는 단어가 들어가는 엑셀 파일을 찾음
for file in os.listdir(folder_path):
    if file.endswith('.xlsx') and "수량-" in file:
        quantity_file_path = os.path.join(folder_path, file)
        break
else:
    print("폴더 내에 '수량-'이라는 단어가 들어가는 엑셀 파일이 없습니다.")
    exit()

# 수량이라는 단어가 들어간 엑셀 파일 경로 출력
print("수량이라는 단어가 들어간 엑셀 파일의 경로:", quantity_file_path)


# 엑셀 파일 내 "용지조서(test)" 시트에서 데이터를 읽어오는 함수
def read_data_from_excel(file_path, sheet_name, start_row, start_col):
    # 엑셀 파일 열기
    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]

    # 시작 셀부터 데이터 읽기
    data = []
    row_index = start_row
    col_index = start_col
    while True:
        cell_value = ws.cell(row=row_index, column=col_index).value
        if cell_value:
            data.append(cell_value)
            row_index += 1
        else:
            break

    # 엑셀 파일 닫기
    wb.close()
    return data


# "용지조서(test)" 시트의 데이터를 읽어옵니다.
data = read_data_from_excel(quantity_file_path, "용지조서(test)", start_row=5, start_col=2)

번지수 = []
for value in data:
    print(value)
    번지수.append(value)

# 단어들을 담은 리스트 출력
print("출력한 단어들을 리스트에 담았습니다:", 번지수)
